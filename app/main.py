import os
import sys
from functools import partial

import openpyxl
from PySide6.QtCore import QTimer, Qt
from PySide6.QtGui import QAction, QColor, QIcon
from PySide6.QtWidgets import (
    QApplication,
    QCheckBox,
    QComboBox,
    QDialog,
    QFileDialog,
    QFrame,
    QHeaderView,
    QHBoxLayout,
    QLabel,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QSplitter,
    QStatusBar,
    QStyle,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QVBoxLayout,
    QWidget,
    QLineEdit,
)

from config import (
    CONFIG_FILE,
    DEFAULT_DST,
    DEFAULT_SRC,
    load_previous_config,
    save_config,
)
from ui.dialogs.pdf_config_dialog import DialogPdfConfig
from ui.dialogs.dialogs import PathDialog
from utils.log_utils import append_log
from services.pdf_generator import save_images_to_pdf_for_dialog_users


class WrongAnswerManager(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("copycopyWA")
        self.resize(800, 700)

        # 아이콘 설정
        if getattr(sys, "frozen", False):
            base_path = sys._MEIPASS
        else:
            base_path = os.path.abspath(".")
        icon_path = os.path.join(base_path, "resources", "icon.ico")
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))

        # 설정 로드
        self.config = load_previous_config() or {"users": []}
        self.users = self.config.get("users", [])
        self.modified_rows = set()
        self.modified = False
        self.is_first_update = True

        # --- 메뉴, 상태바, 타이머 ---
        self.setup_menus()
        self.setup_status_bar()
        self.setup_search_timer()

        # --- 중앙 위젯 ---
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        # --- 버튼 ---
        self.setup_buttons(main_layout)

        # --- 검색 ---
        self.setup_search_ui(main_layout)

        # --- 메인 분할 스플리터 (테이블 영역 & 로그 영역 구분) ---
        splitter = QSplitter(Qt.Vertical)

        # 1) 상단 테이블 영역
        table_container = QWidget()
        table_layout = QVBoxLayout(table_container)
        table_layout.setContentsMargins(0, 0, 0, 0)

        self.select_all_checkbox = QCheckBox("전체 선택")
        self.select_all_checkbox.setToolTip("보이는 모든 항목을 선택/해제합니다.")
        checkbox_layout = QHBoxLayout()
        checkbox_layout.addWidget(self.select_all_checkbox)
        checkbox_layout.addStretch()
        checkbox_layout.setContentsMargins(10, 0, 0, 5)
        table_layout.addLayout(checkbox_layout)

        self.table = QTableWidget(0, 4)
        self.setup_table()
        table_layout.addWidget(self.table)

        splitter.addWidget(table_container)

        # 2) 하단 로그 영역
        log_container = QWidget()
        log_layout = QVBoxLayout(log_container)
        log_layout.setContentsMargins(0, 0, 0, 0)

        log_header_layout = QHBoxLayout()
        log_header_layout.addWidget(QLabel("로그:"))
        log_header_layout.addStretch()
        self.clear_log_btn = QPushButton()
        self.clear_log_btn.setIcon(self.style().standardIcon(QStyle.SP_TrashIcon))
        self.clear_log_btn.setFixedSize(24, 24)
        self.clear_log_btn.setToolTip("로그 지우기")
        log_header_layout.addWidget(self.clear_log_btn)
        log_layout.addLayout(log_header_layout)

        self.log_output = QTextEdit()
        self.log_output.setReadOnly(True)
        log_layout.addWidget(self.log_output)

        splitter.addWidget(log_container)

        # 스플리터 비율 및 초기 크기 설정 (테이블 넓게, 로그 영역 작게)
        splitter.setStretchFactor(0, 4)
        splitter.setStretchFactor(1, 1)
        splitter.setSizes([450, 120])

        main_layout.addWidget(splitter)

        # --- 이벤트 연결 ---
        self.connect_signals()

        # --- 초기화 ---
        self.load_table()
        self.log("🟢 프로그램 시작")
        self.initialize_config()

    def setup_menus(self):
        menu_bar = self.menuBar()
        settings_menu = menu_bar.addMenu("설정")
        path_config_action = QAction("경로 설정 열기", self)
        pdf_config_action = QAction("PDF 설정", self)
        close_action = QAction("닫기", self)
        settings_menu.addAction(path_config_action)
        settings_menu.addAction(pdf_config_action)
        settings_menu.addSeparator()
        settings_menu.addAction(close_action)
        path_config_action.triggered.connect(self.open_path_dialog)
        pdf_config_action.triggered.connect(self.open_config_dialog)
        close_action.triggered.connect(self.close)
        info_menu = menu_bar.addMenu("정보")
        info_action = QAction("버전확인", self)
        info_menu.addAction(info_action)
        info_action.triggered.connect(self.show_version_dialog)

    def setup_status_bar(self):
        self.setStatusBar(QStatusBar(self))
        self.row_count_label = QLabel("총 0개")
        self.statusBar().addPermanentWidget(self.row_count_label)

    def setup_search_timer(self):
        self.search_timer = QTimer(self)
        self.search_timer.setSingleShot(True)
        self.search_timer.timeout.connect(self.filter_table)

    def setup_buttons(self, parent_layout):
        btn_layout_1 = QHBoxLayout()
        self.add_btn = QPushButton("학생 추가")
        self.del_btn = QPushButton("학생 삭제")
        self.save_btn = QPushButton("학생 저장")
        self.pdf_btn = QPushButton("오답노트 PDF 저장")
        self.refresh_btn = QPushButton("새로고침")
        for btn in [self.add_btn, self.del_btn, self.save_btn, self.pdf_btn, self.refresh_btn]:
            btn_layout_1.addWidget(btn)
        parent_layout.addLayout(btn_layout_1)

        separator = QFrame()
        separator.setFrameShape(QFrame.HLine)
        separator.setFrameShadow(QFrame.Sunken)
        parent_layout.addWidget(separator)

        btn_layout_2 = QHBoxLayout()
        btn_layout_2.addWidget(QLabel("학생 관리:"))
        self.excel_export_btn = QPushButton("학생데이터 내보내기")
        self.excel_import_btn = QPushButton("학생데이터 불러오기")
        btn_layout_2.addWidget(self.excel_export_btn)
        btn_layout_2.addWidget(self.excel_import_btn)
        btn_layout_2.addStretch()
        parent_layout.addLayout(btn_layout_2)

    def setup_search_ui(self, parent_layout):
        search_layout = QHBoxLayout()
        self.search_column_combo = QComboBox()
        self.search_column_combo.addItems(["전체", "이름", "오답노트 제목", "오답노트 번호"])
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("검색어 입력 후 Enter 또는 잠시 대기...")
        search_layout.addWidget(QLabel("검색:"))
        search_layout.addWidget(self.search_column_combo)
        search_layout.addWidget(self.search_input)
        parent_layout.addLayout(search_layout)

    def setup_table(self):
        self.table.setHorizontalHeaderLabels(["", "이름", "오답노트 제목", "오답노트 번호"])
        self.table.verticalHeader().setVisible(False)
        self.table.setSortingEnabled(True)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)
        self.table.horizontalHeader().resizeSection(0, 40)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setSelectionMode(QTableWidget.SingleSelection)

    def connect_signals(self):
        self.add_btn.clicked.connect(self.add_row)
        self.del_btn.clicked.connect(self.delete_selected)
        self.save_btn.clicked.connect(self.save_all)
        self.pdf_btn.clicked.connect(self.export_pdf)
        self.refresh_btn.clicked.connect(lambda: self.load_table())
        self.excel_export_btn.clicked.connect(self.export_excel)
        self.excel_import_btn.clicked.connect(self.import_excel)
        self.clear_log_btn.clicked.connect(self.log_output.clear)
        self.table.cellChanged.connect(self.on_cell_changed)
        self.table.cellClicked.connect(self.on_cell_click_row)
        self.search_input.returnPressed.connect(self.filter_table)
        self.search_input.textChanged.connect(lambda: self.search_timer.start(100))
        self.search_column_combo.currentIndexChanged.connect(self.filter_table)
        self.select_all_checkbox.stateChanged.connect(self.toggle_all_checkboxes)

    def log(self, message):
        append_log(self.log_output, message)

    def update_row_count(self):
        total_rows = self.table.rowCount()
        visible_rows = sum(1 for r in range(total_rows) if not self.table.isRowHidden(r))

        search_text = self.search_input.text()
        log_this_update = False

        if search_text:
            text = f"검색: {visible_rows} / 총: {total_rows}개"
            log_this_update = True
        else:
            text = f"총: {total_rows}개"
            if self.is_first_update:
                log_this_update = True

        self.row_count_label.setText(text)
        if log_this_update:
            self.log(f"검색어: '{self.search_input.text()}' 로 검색 : {text}")

        self.is_first_update = False

    # -------------------- 체크박스 관리 --------------------
    def toggle_all_checkboxes(self, state):
        # 사용자가 '전체 선택'을 클릭했을 때만 작동 (프로그램에 의한 변경은 무시)
        if self.select_all_checkbox.isTristate():
            return

        check_state = Qt.CheckState(state)
        for row in range(self.table.rowCount()):
            # 보이는 행에 대해서만 체크박스 상태 변경
            if not self.table.isRowHidden(row):
                widget = self.table.cellWidget(row, 0)
                if widget:
                    checkbox = widget.findChild(QCheckBox)
                    if checkbox:
                        checkbox.setCheckState(check_state)

    # 전체체크
    def update_select_all_state(self):
        visible_rows = [r for r in range(self.table.rowCount()) if not self.table.isRowHidden(r)]
        if not visible_rows:
            self.select_all_checkbox.setCheckState(Qt.Unchecked)
            return

        checked_count = 0
        for row in visible_rows:
            widget = self.table.cellWidget(row, 0)
            if widget:
                checkbox = widget.findChild(QCheckBox)
                if checkbox and checkbox.isChecked():
                    checked_count += 1

        # 시그널 루프 방지를 위해 상태 변경 전 시그널 블락
        self.select_all_checkbox.blockSignals(True)
        if checked_count == 0:
            self.select_all_checkbox.setTristate(False)
            self.select_all_checkbox.setCheckState(Qt.Unchecked)
        elif checked_count == len(visible_rows):
            self.select_all_checkbox.setTristate(False)
            self.select_all_checkbox.setCheckState(Qt.Checked)
        else:
            self.select_all_checkbox.setTristate(True)
            self.select_all_checkbox.setCheckState(Qt.PartiallyChecked)
        self.select_all_checkbox.blockSignals(False)

    # -------------------- 검색 --------------------
    def filter_table(self):
        search_text = self.search_input.text().lower()
        search_column_index = self.search_column_combo.currentIndex()

        for row in range(self.table.rowCount()):
            match = False
            if search_column_index == 0:  # 전체
                texts_to_check = [
                    self.table.item(row, 1).text(),
                    self.table.item(row, 2).text(),
                    self.table.cellWidget(row, 3).findChild(QLineEdit).text(),
                ]
                if any(search_text in t.lower() for t in texts_to_check):
                    match = True
            else:  # 특정 컬럼
                col = search_column_index
                if col == 3:  # 오답노트 번호 (위젯)
                    cell_text = self.table.cellWidget(row, col).findChild(QLineEdit).text()
                else:  # 이름, 제목 (아이템)
                    cell_text = self.table.item(row, col).text()

                if search_text in cell_text.lower():
                    match = True

            self.table.setRowHidden(row, not match)

        self.update_row_count()
        self.update_select_all_state()  # 필터링 후 전체 선택 체크박스 상태 업데이트

    # -------------------- 테이블 관리 --------------------
    def load_table(self, from_config=True):
        if from_config:
            self.search_input.clear()
            self.search_timer.stop()  # 혹시 모를 타이머 중지
            self.filter_table()  # 필터 즉시 초기화

            self.config = load_previous_config() or {"users": []}
            self.users = self.config.get("users", []) if self.config else []
            self.log("🔄 데이터를 새로고침했습니다.")

        self.table.blockSignals(True)
        self.table.setRowCount(0)
        for user in self.users:
            self.add_table_row(user)
        self.table.blockSignals(False)

        self.clear_modified_marks()
        self.filter_table()  # 필터 적용
        self.update_row_count()
        self.update_select_all_state()  # 테이블 로드 후 전체 선택 체크박스 상태 업데이트

    def add_table_row(self, user=None):
        row = self.table.rowCount()
        self.table.insertRow(row)

        checkbox = QCheckBox()
        checkbox.stateChanged.connect(self.update_select_all_state)  # 개별 체크박스 시그널 연결
        w = QWidget()
        l = QHBoxLayout(w)
        l.addWidget(checkbox)
        l.setAlignment(Qt.AlignCenter)
        l.setContentsMargins(0, 0, 0, 0)
        self.table.setCellWidget(row, 0, w)

        name = user.get("name", "") if user else ""
        title = user.get("note_title", "") if user else ""
        numbers = user.get("note_numbers", "") if user else ""

        self.table.setItem(row, 1, QTableWidgetItem(name))
        self.table.setItem(row, 2, QTableWidgetItem(title))

        note_widget = self.create_note_number_widget(numbers, row)
        self.table.setCellWidget(row, 3, note_widget)

    def create_note_number_widget(self, text, row):
        widget = QWidget()
        layout = QHBoxLayout(widget)
        layout.setContentsMargins(2, 0, 2, 0)
        line_edit = QLineEdit(text)
        line_edit.textChanged.connect(lambda: self.on_cell_changed(row, 3))
        button = QPushButton()
        button.setIcon(self.style().standardIcon(QStyle.SP_DirIcon))
        button.setFixedSize(24, 24)
        button.clicked.connect(partial(self.select_note_images_for_row, row))
        layout.addWidget(line_edit)
        layout.addWidget(button)
        widget.setLayout(layout)
        return widget

    def select_note_images_for_row(self, row):
        self.config = load_previous_config()
        folder = self.config.get("source_dir", DEFAULT_SRC)
        if not os.path.exists(folder):
            self.log(f"⚠️ 원본 폴더가 존재하지 않습니다: {folder}")
            QMessageBox.warning(self, "경고", "설정에서 원본 폴더를 먼저 지정해주세요.")
            return

        files, _ = QFileDialog.getOpenFileNames(self, "이미지 파일 선택", folder, "Image Files (*.png *.jpg *.bmp)")
        if files:
            names = [os.path.splitext(os.path.basename(f))[0] for f in files]
            widget = self.table.cellWidget(row, 3)
            line_edit = widget.findChild(QLineEdit)
            line_edit.setText(", ".join(names))
            self.log(f"🟢 {len(names)}개 파일 선택됨 (행: {row + 1})")

    def on_cell_changed(self, row, col):
        self.modified = True
        self.modified_rows.add(row)
        self.mark_row_as_modified(row)

    def mark_row_as_modified(self, row):
        for col in range(1, 3):
            item = self.table.item(row, col)
            if item:
                item.setBackground(QColor(255, 255, 200))
        widget = self.table.cellWidget(row, 3)
        if widget:
            widget.setStyleSheet("background-color: #FFFFC8;")

    def clear_modified_marks(self):
        for row in list(self.modified_rows):
            if row < self.table.rowCount():
                for col in range(1, 3):
                    item = self.table.item(row, col)
                    if item:
                        item.setBackground(QColor(255, 255, 255))
                widget = self.table.cellWidget(row, 3)
                if widget:
                    widget.setStyleSheet("")
        self.modified_rows.clear()
        self.modified = False

    def on_cell_click_row(self, row, col):
        # col == 0 인 체크박스 열만 클릭에 반응하도록 수정
        # QCheckBox 위젯은 자체적으로 클릭 이벤트를 처리하므로,
        # 다른 열을 클릭했을 때 체크박스 상태를 변경하는 기존 로직을 제거합니다.
        if col == 0:
            pass

    # -------------------- 데이터 관리 --------------------
    def add_row(self):
        self.add_table_row()
        self.modified = True
        row = self.table.rowCount() - 1
        self.mark_row_as_modified(row)
        self.update_row_count()
        self.log("➕ 새 행이 추가되었습니다.")
        self.update_select_all_state()

    def delete_selected(self):
        rows_to_remove = [
            r for r in range(self.table.rowCount()) if self.table.cellWidget(r, 0).findChild(QCheckBox).isChecked()
        ]
        if not rows_to_remove:
            QMessageBox.information(self, "안내", "삭제할 행을 선택하세요.")
            return

        for row in sorted(rows_to_remove, reverse=True):
            self.table.removeRow(row)
        self.modified = True
        self.log(f"🗑️ {len(rows_to_remove)}개 행 삭제됨")
        self.update_row_count()
        self.update_select_all_state()

    def save_all(self, silent=False):
        new_users = []
        for row in range(self.table.rowCount()):
            name_item = self.table.item(row, 1)
            name = name_item.text().strip() if name_item else ""
            if name:
                title_item = self.table.item(row, 2)
                title = title_item.text().strip() if title_item else ""
                note_widget = self.table.cellWidget(row, 3)
                numbers = note_widget.findChild(QLineEdit).text().strip()
                new_users.append({"name": name, "note_title": title, "note_numbers": numbers})

        self.users = new_users
        self.config["users"] = self.users
        save_config(self.config)
        self.clear_modified_marks()
        self.log("💾 모든 변경사항이 저장되었습니다.")
        if not silent:
            QMessageBox.information(self, "저장 완료", "모든 변경사항이 저장되었습니다.")

    def export_excel(self):
        file_path, _ = QFileDialog.getSaveFileName(self, "엑셀 파일로 저장", "", "Excel Files (*.xlsx)")
        if not file_path:
            return
        try:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "사용자 데이터"
            headers = ["이름", "오답노트 제목", "오답노트 번호"]
            sheet.append(headers)
            for row in range(self.table.rowCount()):
                name = self.table.item(row, 1).text()
                title = self.table.item(row, 2).text()
                numbers = self.table.cellWidget(row, 3).findChild(QLineEdit).text()
                sheet.append([name, title, numbers])
            wb.save(file_path)
            self.log(f"📄 엑셀 파일로 저장 완료: {file_path}")
            QMessageBox.information(self, "저장 완료", "엑셀 파일로 성공적으로 저장했습니다.")
        except Exception as e:
            self.log(f"❌ 엑셀 저장 실패: {e}")
            QMessageBox.critical(self, "오류", f"엑셀 저장 중 오류가 발생했습니다: {e}")

    def import_excel(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "엑셀 파일 선택", "", "Excel Files (*.xlsx *.xls)")
        if not file_path:
            return
        try:
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
            imported_users = [
                {"name": str(row[0]), "note_title": str(row[1] or ""), "note_numbers": str(row[2] or "")}
                for i, row in enumerate(sheet.iter_rows(values_only=True))
                if i > 0 and row[0]
            ]
            if not imported_users:
                QMessageBox.warning(self, "안내", "엑셀에서 사용자 데이터를 찾지 못했습니다.")
                return
            reply = QMessageBox.question(
                self,
                "데이터 가져오기",
                "기존 데이터를 유지하고 엑셀 데이터를 추가하시겠습니까?\n('아니오'를 선택하면 기존 데이터가 삭제됩니다.)",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes,
            )
            if reply == QMessageBox.No:
                self.users = imported_users
            else:
                existing_names = {u["name"] for u in self.users}
                for u in imported_users:
                    if u["name"] not in existing_names:
                        self.users.append(u)
            self.load_table(from_config=False)
            self.log(f"📊 {len(imported_users)}명의 사용자를 엑셀에서 불러왔습니다.")
            QMessageBox.information(self, "완료", f"{len(imported_users)}명의 사용자를 불러왔습니다.")
        except Exception as e:
            self.log(f"❌ 엑셀 불러오기 실패: {e}")
            QMessageBox.critical(self, "오류", f"엑셀 불러오기 실패: {e}")

    def export_pdf(self):
        if self.modified:
            self.save_all(silent=True)
            self.log("💾 PDF 저장을 위해 변경사항을 자동으로 저장했습니다.")
        checked_users = []
        for row in range(self.table.rowCount()):
            if self.table.cellWidget(row, 0).findChild(QCheckBox).isChecked():
                name_item = self.table.item(row, 1)
                name = name_item.text() if name_item else ""
                title_item = self.table.item(row, 2)
                title = title_item.text() if title_item else ""
                note_widget = self.table.cellWidget(row, 3)
                numbers_text = note_widget.findChild(QLineEdit).text()
                numbers = [num.strip() for num in numbers_text.split(",") if num.strip()]
                if name and numbers:
                    checked_users.append({"name": name, "note_title": title, "note_numbers": numbers})
        if not checked_users:
            QMessageBox.warning(self, "알림", "PDF로 저장할 사용자를 선택하세요.")
            return
        self.config = load_previous_config()
        save_images_to_pdf_for_dialog_users(
            config=self.config,
            users=checked_users,
            log_callback=self.log,
            parent_widget=self,
        )

    # -------------------- 메뉴 액션 및 기타 --------------------
    def open_path_dialog(self):
        dlg = PathDialog(self)
        if dlg.exec():
            self.log("경로 설정 저장 완료!")

    def open_config_dialog(self):
        dialog = DialogPdfConfig(self)
        if dialog.exec():
            self.log("PDF 설정 완료!")

    def initialize_config(self):
        if not os.path.exists(CONFIG_FILE) or not self.config.get("configured", False):
            # 1. 최초 실행 환영 안내 메시지
            QMessageBox.information(
                self,
                "최초 실행 안내 (온보딩 가이드)",
                "오답노트 관리 프로그램에 오신 것을 환영합니다!\n\n"
                "처음 실행하시는 사용자를 위해 필수 설정 2단계를 진행합니다:\n\n"
                " [1단계: 경로 설정] - PDF 템플릿 파일, 이미지 원본 폴더(오답노트로 사용할 문제 이미지들이 들어있는 폴더), PDF 저장 대상 폴더 지정\n"
                " [2단계: PDF 이미지 설정] - PDF 페이지 내 이미지 크기, 여백 및 위치 조정\n\n"
                "'확인' 버튼을 누르시면 1단계 [경로 설정]으로 이동합니다.",
            )

            # --- 1단계: 경로 설정 (PathDialog) ---
            path_dlg = PathDialog(self)
            path_dlg.setWindowTitle("경로 설정 (1/2단계)")
            if path_dlg.exec() == QDialog.Accepted:
                self.log("🟢 1단계: 경로 설정이 저장되었습니다.")
            else:
                self.log("⚠️ 1단계: 경로 설정을 건너뛰었습니다.")

            # 2단계 안내 메시지
            QMessageBox.information(
                self,
                "2단계: PDF 이미지 및 여백 설정 안내",
                "이어서 [2단계: PDF 이미지 및 여백 설정]을 진행합니다.\n\n"
                "오답노트 PDF 출력 시 삽입될 이미지의 가로/세로 크기 및 좌우/상하 여백을 조정할 수 있습니다.\n"
                "우측 미리보기 화면을 참고하여 설정해 주세요.(추후에 수정할수있습니다.)",
            )

            # --- 2단계: PDF 이미지 설정 (DialogPdfConfig) ---
            pdf_dlg = DialogPdfConfig(self)
            pdf_dlg.setWindowTitle("PDF 이미지 설정 (2/2단계)")
            if pdf_dlg.exec() == QDialog.Accepted:
                self.log("🟢 2단계: PDF 이미지 설정이 저장되었습니다.")
            else:
                self.log("⚠️ 2단계: PDF 이미지 설정을 건너뛰었습니다.")

            # 설정 완료 상태 저장
            self.config = load_previous_config() or {}
            self.config["configured"] = True
            save_config(self.config)
            self.log("🎉 모든 초기 설정 단계가 완료되었습니다.")

            # 완료 안내 메시지
            QMessageBox.information(
                self,
                "초기 설정 완료",
                "🎉 축하합니다! 모든 초기 설정 과정이 완료되었습니다.\n\n"
                "이제 학생 데이터를 등록하고 '오답노트 PDF 저장' 버튼을 누르면 설정된 양식대로 PDF가 자동 생성됩니다.\n"
                "(설정은 언제든지 메뉴의 '설정' 항목에서 수정할 수 있습니다.)",
            )

    def show_version_dialog(self):
        QMessageBox.information(self, "버전 정보", "오답노트 관리 프로그램 v1.1.0")

    def closeEvent(self, event):
        if self.modified:
            reply = QMessageBox.question(
                self,
                "저장되지 않은 변경사항",
                "저장하지 않은 변경사항이 있습니다. 저장하시겠습니까?",
                QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel,
                QMessageBox.Cancel,
            )
            if reply == QMessageBox.Yes:
                self.save_all(silent=True)
                event.accept()
            elif reply == QMessageBox.No:
                event.accept()
            else:
                event.ignore()
        else:
            event.accept()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = WrongAnswerManager()
    window.show()
    sys.exit(app.exec())
